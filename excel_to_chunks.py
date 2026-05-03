"""
excel_to_chunks.py
==================
Universal parser: reads ALL sheets from any Excel file.
Column headers are taken from `header_row`; data starts at `data_start_row`.
Each non-empty data row becomes one chunk: "Column: value\n..."

Settings: config.yaml
Usage:
    python excel_to_chunks.py
"""

import json
from pathlib import Path

import openpyxl
import yaml


def load_config(path: str = "config.yaml") -> dict:
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def val(cell) -> str:
    v = cell.value
    return str(v).strip() if v is not None else ""


def is_placeholder(text: str, skip_values: list[str]) -> bool:
    t = text.strip().upper()
    return any(t == s.strip().upper() for s in skip_values)


def make_chunk(text: str, sheet: str, row_idx: int = 0) -> dict | None:
    text = text.strip()
    if not text:
        return None
    return {
        "text": text,
        "metadata": {"sheet": sheet, "row": row_idx},
    }


def parse_sheet(ws, header_row: int, data_start_row: int, skip_values: list[str]) -> list[dict]:
    if ws.max_row is None or ws.max_row < header_row:
        return []

    try:
        header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row))
    except StopIteration:
        return []

    headers = [val(c) for c in header_cells]
    if not any(headers):
        return []

    chunks = []
    for row in ws.iter_rows(min_row=data_start_row):
        parts = []
        for header, cell in zip(headers, row):
            if not header:
                continue
            v = val(cell)
            if not v or is_placeholder(v, skip_values):
                continue
            parts.append(f"{header}: {v}")

        if parts:
            chunk = make_chunk("\n".join(parts), ws.title, row[0].row)
            if chunk:
                chunks.append(chunk)

    return chunks


def extract_chunks(config: dict) -> list[dict]:
    excel_path = config["excel"]["path"]
    header_row = config["excel"].get("header_row", 1)
    data_start_row = config["excel"].get("data_start_row", 2)
    skip_sheets = set(config["excel"].get("skip_sheets", []))
    skip_values = config["excel"].get("skip_values", ["FILL"])

    if not Path(excel_path).exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    try:
        all_chunks: list[dict] = []
        for sheet_name in wb.sheetnames:
            if sheet_name in skip_sheets:
                print(f"  ⏭️  Skipping: '{sheet_name}'")
                continue
            chunks = parse_sheet(wb[sheet_name], header_row, data_start_row, skip_values)
            all_chunks.extend(chunks)
            print(f"  ✅ {sheet_name}: {len(chunks)} chunks")
    finally:
        wb.close()

    return all_chunks


if __name__ == "__main__":
    cfg = load_config()
    print(f"\n📖 Parsing: {cfg['excel']['path']}")

    chunks = extract_chunks(cfg)
    print(f"\n📦 Total chunks: {len(chunks)}")

    out_path = cfg["output"]["chunks_file"]
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(chunks, f, ensure_ascii=False, indent=2)
    print(f"💾 Saved to {out_path}")

    print("\n── Preview (first 2 chunks) ──")
    for ch in chunks[:2]:
        print(f"\n[{ch['metadata']['sheet']}]\n{ch['text'][:300]}")
        print("─" * 50)
